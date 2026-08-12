#!/usr/bin/env python3
"""
Extrait les effectifs PASA depuis l'export RenoiRH ETPT_RPROG
(feuille « Données annuelles ») vers public/data/pasa-effectifs.json.

Usage :
  python3 scripts/build_pasa_effectifs.py
  python3 scripts/build_pasa_effectifs.py trdata/2026_03_Suivi_des_emplois_en_ETPT_RPROG.xlsx
"""

from __future__ import annotations

import json
import sys
from collections import defaultdict
from datetime import datetime
from pathlib import Path
from typing import Any, Optional

import openpyxl

SHEET_NAME = "Données annuelles"
HEADER_ROW = 5

COL = {
    "matricule": "Etat Civil : Matricule SIRH",
    "mois": "Mois observation (synth annee)",
    "niveau03": "Niveau 03 Opé. : Libellé Court",
    "niveau06": "Niveau 06 Opé. : Libellé Court",
    "niveau08": "Niveau 08 Opé. : Libellé Court",
    "poste": "ETPT RH : Poste Libellé long",
    "sous_action": "ETPT RH : Sous-Action",
    "etpt": "ETPT RH",
}

# ---------------------------------------------------------------------------
# Configuration de tous les PASA
# ---------------------------------------------------------------------------

PASA_CONFIG: list[dict[str, Any]] = [
    {
        "id": "pasa2",
        "code": "0217-11-02",
        "title": "PASA 2 — Emplois et formations maritimes",
        "categories_order": [
            "lpm", "services_formation", "services_instructions", "ssgm",
            "gens_de_mer_ac", "autres",
        ],
    },
    {
        "id": "pasa3",
        "code": "0217-11-03",
        "title": "PASA 3 — Flotte de commerce et sécurité des navires",
        "categories_order": [
            "flotte_commerce_ac", "securite_maritime_ac", "autres",
        ],
    },
    {
        "id": "pasa4",
        "code": "0217-11-04",
        "title": "PASA 4 — Contrôle des activités en mer",
        "categories_order": ["controle_ac", "autres"],
    },
    {
        "id": "pasa5",
        "code": "0217-11-05",
        "title": "PASA 5 — Soutien",
        "categories_order": ["tous"],
    },
    {
        "id": "pasa7",
        "code": "0217-11-07",
        "title": "PASA 7 — Pêche et aquaculture",
        "categories_order": [
            "peche", "aquaculture", "peche_aqua_ac", "autres",
        ],
    },
    {
        "id": "pasa8",
        "code": "0217-11-08",
        "title": "PASA 8 — Planification et plaisance",
        "categories_order": [
            "planification", "plaisance",
            "plaisance_ac", "planification_ac",
            "autres",
        ],
    },
    {
        "id": "pasa11",
        "code": "0217-11-11",
        "title": "PASA 11 — CROSS",
        "categories_order": ["cross_ac", "autres"],
    },
    {
        "id": "pasa13",
        "code": "0217-11-13",
        "title": "PASA 13 — Phares et Balises (dont POLMAR)",
        "categories_order": ["phares_balises_ac", "autres"],
    },
    {
        "id": "pasa16",
        "code": "0217-11-16",
        "title": "PASA 16 — Capitaineries",
        "categories_order": ["capitaineries_ac", "autres"],
    },
]

LABELS = {
    "lpm": "Lycées professionnels maritimes (LPM)",
    "services_formation": "Services formation (DIRM)",
    "services_instructions": "Services instructions / permis (DDTM)",
    "ssgm": "Services de santé des gens de mer (SSGM)",
    "gens_de_mer_ac": "Gens de mer administration centrale",
    "flotte_commerce_ac": "Flotte de commerce administration centrale",
    "securite_maritime_ac": "Sécurité maritime administration centrale",
    "controle_ac": "Contrôle des activités en mer administration centrale",
    "cross_ac": "CROSS administration centrale",
    "phares_balises_ac": "Phares et Balises administration centrale",
    "capitaineries_ac": "Capitaineries administration centrale",
    "planification": "Planification (DIRM)",
    "plaisance": "Plaisance (DDTM)",
    "planification_ac": "Planification administration centrale",
    "plaisance_ac": "Plaisance administration centrale",
    "peche": "Pêche",
    "aquaculture": "Aquaculture",
    "peche_aqua_ac": "Pêche et aquaculture administration centrale",
    "tous": "Tous les agents",
    "autres": "Autres / non classés",
}


def _txt(value: Any) -> str:
    if value is None:
        return ""
    return str(value).strip()


def _float(value: Any) -> float:
    try:
        if value is None or value == "":
            return 0.0
        return float(value)
    except (TypeError, ValueError):
        return 0.0


def is_dirm(service: str) -> bool:
    s = service.upper()
    return s.startswith("DIRM") or s == "ESP MER"


def is_ddtm(service: str) -> bool:
    s = service.upper()
    return s.startswith("DDTM") or s.startswith("DML")


def is_dgampa(service: str) -> bool:
    return service.upper().strip() == "DGAMPA"


def contains_any(text: str, needles: tuple[str, ...]) -> bool:
    up = text.upper()
    return any(n.upper() in up for n in needles)


def n08_starts_with(n08: str, prefix: str) -> bool:
    return n08.upper().strip().startswith(prefix.upper())


# ---------------------------------------------------------------------------
# Fonctions de catégorisation par PASA
# ---------------------------------------------------------------------------

PASA2_INSTRUCTION_KEYWORDS = ("ACTIVITÉS MARITIMES", "INSTRUCT")

def categorize_pasa2(r: dict[str, Any]) -> str:
    n03, n06, n08, poste = r["niveau03"], r["niveau06"], r["niveau08"], r["poste"]
    if is_dgampa(n03):
        return "gens_de_mer_ac"
    if contains_any(n06, ("LPM",)) or contains_any(n08, ("LPM",)):
        return "lpm"
    if contains_any(n06, ("SSGM",)) or contains_any(n08, ("SSGM",)):
        return "ssgm"
    if is_dirm(n03):
        return "services_formation"
    if is_ddtm(n03):
        return "services_instructions"
    if contains_any(poste, PASA2_INSTRUCTION_KEYWORDS):
        return "services_instructions"
    return "autres"


PASA8_PLANIFICATION_KW = (
    "DOMAINE PUBLIC", "ÉCONOMIE BLEUE", "ECONOMIE BLEUE", "DPM",
    "COORDINATION", "POLITIQUES PUBLIQUES", "AFFAIRES ÉCONOMIQUES",
    "AFFAIRES ECONOMIQUES", "ESPACES MARITIMES",
    "ÉCONOMIE", "ECONOMIE", "LITT.", "MARITIMES", "SDDM",
)
PASA8_PLAISANCE_KW = (
    "PLAISANCE", "ACTIVITÉS MARITIMES", "ACTIVITES MARITIMES",
    "PROTECTION SOCIALE", "MARIN", "GENS",
)

def categorize_pasa8(r: dict[str, Any]) -> str:
    n03, n08, poste = r["niveau03"], r["niveau08"], r["poste"]
    if is_dgampa(n03):
        if n08_starts_with(n08, "SEML/MNP"):
            return "plaisance_ac"
        if n08_starts_with(n08, "SEML/PM"):
            return "planification_ac"
        return "planification_ac"
    if is_dirm(n03):
        return "planification"
    if is_ddtm(n03):
        return "plaisance"
    if contains_any(poste, PASA8_PLANIFICATION_KW):
        return "planification"
    if contains_any(poste, PASA8_PLAISANCE_KW):
        return "plaisance"
    return "autres"


PASA7_PECHE_KW = (
    "RÉGLEMENTATION", "REGLEMENTATION", "FEAMP", "RESSOURCES HALIEUTIQUES",
    "FILIÈRES", "FILIERES", "CONTRÔLE", "CONTROLE", "ÉCONOMIE", "ECONOMIE",
    "CAPTURES",
    "TERRITORIALE", "CHEF", "MER", "LITTORAL", "ÉCONOMIQUE", "ECONOMIQUE",
    "SGMPC", "AFFAIRES", "MARITIMES",
)
PASA7_AQUA_KW = (
    "CONCHYLICOLE", "CULT. MARINES", "ALGOCULTURE", "URH",
    "TECHNICIEN", "CARTOGRAPHIE",
)
PASA7_PECHE_LEGACY = ("PÊCHE", "PECHE", "PECHERIE", "/BEP", "/BGR", "/BASD", "/BAEI", "PECH ")
PASA7_AQUA_LEGACY = ("AQUACULTURE", "AQUA", "CULTURES MARINES", "CULTURE MARINE", "BAQUA")

def categorize_pasa7(r: dict[str, Any]) -> str:
    n06, n08, poste, n03 = r["niveau06"], r["niveau08"], r["poste"], r["niveau03"]
    if is_dgampa(n03):
        return "peche_aqua_ac"
    text = " ".join([poste, n06, n08])
    if contains_any(text, PASA7_AQUA_LEGACY) or contains_any(poste, PASA7_AQUA_KW):
        return "aquaculture"
    if contains_any(text, PASA7_PECHE_LEGACY) or contains_any(poste, PASA7_PECHE_KW):
        return "peche"
    return "autres"


def categorize_pasa3(r: dict[str, Any]) -> str:
    n03, n08 = r["niveau03"], r["niveau08"]
    if is_dgampa(n03):
        if n08_starts_with(n08, "SFM/MFC"):
            return "flotte_commerce_ac"
        if n08_starts_with(n08, "SFM/STEN"):
            return "securite_maritime_ac"
        return "flotte_commerce_ac"
    return "autres"


def categorize_simple_dgampa(ac_key: str) -> Any:
    def _cat(r: dict[str, Any]) -> str:
        if is_dgampa(r["niveau03"]):
            return ac_key
        return "autres"
    return _cat


def categorize_pasa5(r: dict[str, Any]) -> str:
    return "tous"


CATEGORIZERS: dict[str, Any] = {
    "pasa2": categorize_pasa2,
    "pasa3": categorize_pasa3,
    "pasa4": categorize_simple_dgampa("controle_ac"),
    "pasa5": categorize_pasa5,
    "pasa7": categorize_pasa7,
    "pasa8": categorize_pasa8,
    "pasa11": categorize_simple_dgampa("cross_ac"),
    "pasa13": categorize_simple_dgampa("phares_balises_ac"),
    "pasa16": categorize_simple_dgampa("capitaineries_ac"),
}


# ---------------------------------------------------------------------------
# Chargement et agrégation
# ---------------------------------------------------------------------------

def load_rows(chemin: Path) -> tuple[list[dict[str, Any]], int, Optional[int]]:
    wb = openpyxl.load_workbook(chemin, read_only=True, data_only=True)
    if SHEET_NAME not in wb.sheetnames:
        raise SystemExit(f"Feuille « {SHEET_NAME} » absente. Feuilles : {wb.sheetnames}")

    ws = wb[SHEET_NAME]
    rows = list(ws.iter_rows(values_only=True))
    header = rows[HEADER_ROW - 1]
    col_idx = {h: i for i, h in enumerate(header) if h}

    for key, name in COL.items():
        if name not in col_idx:
            raise SystemExit(f"Colonne absente ({name!r}). Colonnes : {[h for h in header if h]}")

    def get(row: tuple[Any, ...], key: str) -> Any:
        idx = col_idx[COL[key]]
        return row[idx] if idx < len(row) else None

    parsed: list[dict[str, Any]] = []
    mois_values: list[int] = []

    for row in rows[HEADER_ROW:]:
        matricule = get(row, "matricule")
        if matricule is None:
            continue
        mois_raw = get(row, "mois")
        try:
            mois = int(float(mois_raw)) if mois_raw is not None else None
        except (TypeError, ValueError):
            mois = None
        if mois is not None:
            mois_values.append(mois)

        parsed.append(
            {
                "matricule": str(matricule).strip(),
                "mois": mois,
                "niveau03": _txt(get(row, "niveau03")),
                "niveau06": _txt(get(row, "niveau06")),
                "niveau08": _txt(get(row, "niveau08")),
                "poste": _txt(get(row, "poste")),
                "sous_action": _txt(get(row, "sous_action")),
                "etpt": _float(get(row, "etpt")),
            }
        )

    mois_ref = max(mois_values) if mois_values else None
    return parsed, len(parsed), mois_ref


def aggregate(rows: list[dict[str, Any]], mois_ref: Optional[int], pasa_cfg: dict[str, Any]) -> dict[str, Any]:
    pasa_key = pasa_cfg["id"]
    code = pasa_cfg["code"]
    categorize = CATEGORIZERS[pasa_key]

    filtered = [
        r for r in rows
        if r["sous_action"].startswith(code) and (mois_ref is None or r["mois"] == mois_ref)
    ]

    effectifs: dict[str, set[str]] = defaultdict(set)
    etpt: dict[str, float] = defaultdict(float)
    details: dict[str, dict[str, dict[str, float]]] = defaultdict(
        lambda: defaultdict(lambda: {"effectif": 0.0, "etpt": 0.0})
    )
    postes: dict[str, dict[tuple[str, str], dict[str, Any]]] = defaultdict(dict)

    for row in filtered:
        if row["etpt"] <= 0:
            continue
        cat = categorize(row)
        effectifs[cat].add(row["matricule"])
        etpt[cat] += row["etpt"]
        svc = row["niveau03"] or "Non renseigné"
        details[cat][svc]["effectif"] += 1
        details[cat][svc]["etpt"] += row["etpt"]

        poste_label = row["poste"] or "Poste non renseigné"
        poste_k = (poste_label, svc)
        if poste_k not in postes[cat]:
            postes[cat][poste_k] = {"matricules": set(), "etpt": 0.0}
        postes[cat][poste_k]["matricules"].add(row["matricule"])
        postes[cat][poste_k]["etpt"] += row["etpt"]

    categories = []
    for key in pasa_cfg["categories_order"]:
        categories.append(
            {
                "id": key,
                "label": LABELS[key],
                "effectif": len(effectifs.get(key, set())),
                "etpt": round(etpt.get(key, 0.0), 2),
                "detailsParService": sorted(
                    [
                        {
                            "service": service,
                            "effectif": int(values["effectif"]),
                            "etpt": round(values["etpt"], 2),
                        }
                        for service, values in details.get(key, {}).items()
                    ],
                    key=lambda x: (-x["etpt"], x["service"]),
                ),
                "detailsParPoste": sorted(
                    [
                        {
                            "poste": poste,
                            "service": service,
                            "effectif": len(values["matricules"]),
                            "etpt": round(values["etpt"], 2),
                        }
                        for (poste, service), values in postes.get(key, {}).items()
                    ],
                    key=lambda x: (-x["etpt"], x["poste"], x["service"]),
                ),
            }
        )

    total_effectif = len({r["matricule"] for r in filtered if r["etpt"] > 0})
    total_etpt = round(sum(r["etpt"] for r in filtered if r["etpt"] > 0), 2)

    return {
        "id": pasa_key,
        "code": code.replace("0217-11-", "217-11-"),
        "title": pasa_cfg["title"],
        "lignesAnalysees": len(filtered),
        "totalEffectif": total_effectif,
        "totalEtpt": total_etpt,
        "categories": categories,
    }


def build(chemin: Path) -> dict[str, Any]:
    rows, nb_brut, mois_ref = load_rows(chemin)
    return {
        "metadonnees": {
            "dateExport": datetime.now().strftime("%Y-%m-%d"),
            "source": chemin.name,
            "feuille": SHEET_NAME,
            "moisReference": mois_ref,
            "lignesBrutes": nb_brut,
            "description": "Effectifs PASA dérivés indépendamment de agents.json",
        },
        "actions": [aggregate(rows, mois_ref, cfg) for cfg in PASA_CONFIG],
    }


def main() -> None:
    base = Path(__file__).parent.parent
    args = [a for a in sys.argv[1:] if str(a).lower().endswith(".xlsx")]
    chemin = Path(args[0]) if args else base / "trdata" / "2026_03_Suivi_des_emplois_en_ETPT_RPROG.xlsx"
    if not chemin.exists():
        raise SystemExit(f"Fichier introuvable : {chemin}")

    data = build(chemin)
    for out in (
        base / "public" / "data" / "pasa-effectifs.json",
        base / "src" / "data" / "pasa-effectifs.json",
    ):
        out.parent.mkdir(parents=True, exist_ok=True)
        with open(out, "w", encoding="utf-8") as f:
            json.dump(data, f, indent=2, ensure_ascii=False)
        print(f"Écrit : {out}")

    for action in data["actions"]:
        print(f"\n{action['title']} (mois {data['metadonnees']['moisReference']})")
        for cat in action["categories"]:
            if cat["effectif"] or cat["etpt"]:
                print(f"  - {cat['label']}: {cat['effectif']} agents, {cat['etpt']} ETPT")


if __name__ == "__main__":
    main()
